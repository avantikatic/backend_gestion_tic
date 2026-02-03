import traceback
from Utils.tools import Tools, CustomException
from Utils.querys import Querys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import date, datetime

class Licencias:

    def __init__(self, db):
        self.db = db
        self.tools = Tools()
        self.querys = Querys(self.db)
        self.token = None

    def crear_licencia(self, data):
        """
        Crea una nueva licencia/servicio en la base de datos
        """
        try:
            
            licencia = self.querys.crear_licencia(data)
            return self.tools.output(201, "Licencia creada exitosamente.", licencia)
                
        except CustomException as ce:
            return self.tools.output(400, str(ce), {})
        except Exception as e:
            print(f"Error creando licencia: {e}")
            return self.tools.output(500, "Error creando licencia.", {})

    def obtener_licencias(self, filtros=None, page=1, per_page=5):
        """
        Obtiene todas las licencias/servicios de la base de datos con paginación
        Permite filtrar por estado (activas/bajas)
        """
        try:
            resultado = self.querys.obtener_licencias(filtros, page, per_page)
            return self.tools.output(200, "Licencias obtenidas exitosamente.", resultado)
                
        except CustomException as ce:
            return self.tools.output(500, str(ce), {})
        except Exception as e:
            print(f"Error obteniendo licencias: {e}")
            return self.tools.output(500, "Error obteniendo licencias.", {})

    def obtener_licencia_por_id(self, licencia_id):
        """
        Obtiene una licencia específica por su ID
        """
        try:
            licencia = self.querys.obtener_licencia_por_id(licencia_id)
            return self.tools.output(200, "Licencia obtenida exitosamente.", licencia)
                
        except CustomException as ce:
            return self.tools.output(404, str(ce), {})
        except Exception as e:
            print(f"Error obteniendo licencia por ID: {e}")
            return self.tools.output(500, "Error obteniendo licencia.", {})

    def actualizar_licencia(self, licencia_id, data):
        """
        Actualiza una licencia existente
        """
        try:
            licencia = self.querys.actualizar_licencia(licencia_id, data)
            return self.tools.output(200, "Licencia actualizada exitosamente.", licencia)
                
        except CustomException as ce:
            return self.tools.output(404, str(ce), {})
        except Exception as e:
            print(f"Error actualizando licencia: {e}")
            return self.tools.output(500, "Error actualizando licencia.", {})

    def eliminar_licencia(self, licencia_id):
        """
        Elimina (marca como inactiva) una licencia
        """
        try:
            self.querys.eliminar_licencia(licencia_id)
            return self.tools.output(200, "Licencia eliminada exitosamente.", {})
                
        except CustomException as ce:
            return self.tools.output(404, str(ce), {})
        except Exception as e:
            print(f"Error eliminando licencia: {e}")
            return self.tools.output(500, "Error eliminando licencia.", {})

    def obtener_historial_licencia(self, licencia_id):
        """
        Obtiene el historial de cambios de una licencia
        """
        try:
            historial = self.querys.obtener_historial_licencia(licencia_id)
            return self.tools.output(200, "Historial obtenido exitosamente.", historial)
                
        except CustomException as ce:
            return self.tools.output(404, str(ce), [])
        except Exception as e:
            print(f"Error obteniendo historial: {e}")
            return self.tools.output(500, "Error obteniendo historial.", [])

    # Función para obtener indicadores de gestión mensual
    def obtener_indicadores_gestion(self, data=None):
        """
        Obtiene los indicadores de gestión por mes:
        - Tickets completados
        - Tickets cerrados oportunamente (fecha_cierre <= fecha_vencimiento)
        - Tickets cerrados no oportunamente (fecha_cierre > fecha_vencimiento)
        - Porcentaje de cumplimiento
        """
        try:
            filtros = data or {}
            anio = filtros.get('anio', datetime.now().year)
            
            # Obtener indicadores usando query
            indicadores = self.querys.obtener_indicadores_gestion(anio)
            
            return self.tools.output(200, "Indicadores de gestión obtenidos exitosamente.", indicadores)
                
        except Exception as e:
            print(f"Error obteniendo indicadores de gestión: {e}")
            return self.tools.output(500, "Error obteniendo indicadores de gestión.", {})

    # ===================================================
    # MÉTODOS PARA CATÁLOGOS/MAESTROS
    # ===================================================

    def obtener_tipos_servicio(self):
        """Obtiene todos los tipos de servicio activos"""
        try:
            tipos = self.querys.obtener_tipos_servicio()
            return self.tools.output(200, "Tipos de servicio obtenidos exitosamente.", tipos)
                
        except CustomException as ce:
            return self.tools.output(500, str(ce), [])
        except Exception as e:
            print(f"Error obteniendo tipos de servicio: {e}")
            return self.tools.output(500, "Error obteniendo tipos de servicio.", [])

    def obtener_proveedores(self):
        """
        Obtiene todos los proveedores activos desde la tabla 'terceros'.
        Retorna lista de objetos con estructura: [{id: nit, nombre: nombres}, ...]
        """
        try:
            proveedores = self.querys.obtener_proveedores()
            return self.tools.output(200, "Proveedores obtenidos exitosamente.", proveedores)
                
        except CustomException as ce:
            return self.tools.output(500, str(ce), [])
        except Exception as e:
            print(f"Error obteniendo proveedores: {e}")
            return self.tools.output(500, "Error obteniendo proveedores.", [])

    def obtener_productos_servicios(self):
        """Obtiene todos los productos/servicios activos"""
        try:
            productos = self.querys.obtener_productos_servicios()
            return self.tools.output(200, "Productos/servicios obtenidos exitosamente.", productos)
                
        except CustomException as ce:
            return self.tools.output(500, str(ce), [])
        except Exception as e:
            print(f"Error obteniendo productos/servicios: {e}")
            return self.tools.output(500, "Error obteniendo productos/servicios.", [])

    def obtener_metodos_pago(self):
        """Obtiene todos los métodos de pago activos"""
        try:
            metodos = self.querys.obtener_metodos_pago()
            return self.tools.output(200, "Métodos de pago obtenidos exitosamente.", metodos)
                
        except CustomException as ce:
            return self.tools.output(500, str(ce), [])
        except Exception as e:
            print(f"Error obteniendo métodos de pago: {e}")
            return self.tools.output(500, "Error obteniendo métodos de pago.", [])

    def obtener_tipos_moneda(self):
        """Obtiene todos los tipos de moneda activos"""
        try:
            tipos = self.querys.obtener_tipos_moneda()
            return self.tools.output(200, "Tipos de moneda obtenidos exitosamente.", tipos)
                
        except CustomException as ce:
            return self.tools.output(500, str(ce), [])
        except Exception as e:
            print(f"Error obteniendo tipos de moneda: {e}")
            return self.tools.output(500, "Error obteniendo tipos de moneda.", [])

    def crear_proveedor(self, nombre):
        """Crea un nuevo proveedor"""
        try:
            nombre_clean = nombre.strip()
            if not nombre_clean:
                return self.tools.output(400, "El nombre del proveedor es requerido.", {})
            
            proveedor = self.querys.crear_proveedor(nombre_clean)
            
            if not proveedor:
                return self.tools.output(500, "Error: No se pudo crear el proveedor.", {})
            
            return self.tools.output(201, "Proveedor creado exitosamente.", proveedor)
                
        except CustomException as ce:
            return self.tools.output(400, str(ce), {})
        except Exception as e:
            print(f"Error creando proveedor: {e}")
            return self.tools.output(500, "Error creando proveedor.", {})

    def crear_producto_servicio(self, nombre):
        """Crea un nuevo producto/servicio"""
        try:
            nombre_clean = nombre.strip()
            if not nombre_clean:
                return self.tools.output(400, "El nombre del producto/servicio es requerido.", {})
            
            producto = self.querys.crear_producto_servicio(nombre_clean)
            
            # Si el producto ya existía, devolver mensaje informativo
            productos_existentes = self.querys.obtener_productos_servicios()
            if any(p['nombre'] == nombre_clean for p in productos_existentes):
                # Si ya existía antes de intentar crearlo
                if producto:
                    return self.tools.output(200, "El producto/servicio ya existe.", producto)
            
            return self.tools.output(201, "Producto/servicio creado exitosamente.", producto)
                
        except CustomException as ce:
            return self.tools.output(400, str(ce), {})
        except Exception as e:
            print(f"Error creando producto/servicio: {e}")
            return self.tools.output(500, "Error creando producto/servicio.", {})

    def crear_tipo_servicio(self, nombre):
        """Crea un nuevo tipo de servicio"""
        try:
            nombre_clean = nombre.strip()
            if not nombre_clean:
                return self.tools.output(400, "El nombre del tipo de servicio es requerido.", {})
            
            tipo_servicio = self.querys.crear_tipo_servicio(nombre_clean)
            
            if not tipo_servicio:
                return self.tools.output(500, "Error: No se pudo crear el tipo de servicio.", {})
            
            return self.tools.output(201, "Tipo de servicio creado exitosamente.", tipo_servicio)
                
        except CustomException as ce:
            return self.tools.output(400, str(ce), {})
        except Exception as e:
            print(f"Error creando tipo de servicio: {e}")
            return self.tools.output(500, "Error creando tipo de servicio.", {})

    def crear_metodo_pago(self, nombre):
        """Crea un nuevo método de pago"""
        try:
            nombre_clean = nombre.strip()
            if not nombre_clean:
                return self.tools.output(400, "El nombre del método de pago es requerido.", {})
            
            metodo_pago = self.querys.crear_metodo_pago(nombre_clean)
            
            if not metodo_pago:
                return self.tools.output(500, "Error: No se pudo crear el método de pago.", {})
            
            return self.tools.output(201, "Método de pago creado exitosamente.", metodo_pago)
                
        except CustomException as ce:
            return self.tools.output(400, str(ce), {})
        except Exception as e:
            print(f"Error creando método de pago: {e}")
            return self.tools.output(500, "Error creando método de pago.", {})




    # ===================================================
    # MÉTODOS PARA REVISIONES GENERALES
    # ===================================================

    def obtener_tipos_revision(self):
        """Obtiene todos los tipos de revisión disponibles"""
        try:
            tipos = self.querys.obtener_tipos_revision()
            return self.tools.output(200, "Tipos de revisión obtenidos exitosamente.", tipos)
        except CustomException as ce:
            return self.tools.output(500, str(ce), [])
        except Exception as e:
            print(f"Error obteniendo tipos de revisión: {e}")
            return self.tools.output(500, "Error obteniendo tipos de revisión.", [])

    def crear_revision(self, data):
        """Crea una nueva revisión general"""
        try:
            revision = self.querys.crear_revision(data)
            return self.tools.output(201, "Revisión creada exitosamente.", revision)
        except CustomException as ce:
            return self.tools.output(400, str(ce), {})
        except Exception as e:
            print(f"Error creando revisión: {e}")
            return self.tools.output(500, "Error creando revisión.", {})

    def obtener_revisiones(self, page=1, per_page=5):
        """Obtiene revisiones con paginación"""
        try:
            resultado = self.querys.obtener_revisiones(page, per_page)
            return self.tools.output(200, "Revisiones obtenidas exitosamente.", resultado)
        except CustomException as ce:
            return self.tools.output(500, str(ce), {})
        except Exception as e:
            print(f"Error obteniendo revisiones: {e}")
            return self.tools.output(500, "Error obteniendo revisiones.", {})

    def eliminar_revision(self, data: dict):
        """Elimina (marca como inactiva) una revisión"""
        try:
            revision_id = data.get('revision_id')
            
            if not revision_id:
                raise CustomException("No se proporcionó revision_id")
            
            resultado = self.querys.eliminar_revision(revision_id)
            return self.tools.output(200, resultado["message"], resultado)
        except CustomException as ce:
            return self.tools.output(404, str(ce), {})
        except Exception as e:
            print(f"Error eliminando revisión: {e}")
            return self.tools.output(500, "Error eliminando revisión.", {})
    
    def exportar_licencias_excel(self, data: dict):
        """Genera un archivo Excel con todas las licencias filtradas"""
        try:
            filtros = data.get('filtros', {})
            
            # Obtener todas las licencias sin paginación
            licencias = self.querys.obtener_todas_licencias_excel(filtros)
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Licencias"
            
            # Estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezados
            headers = [
                'ID', 'Tipo Servicio', 'Proveedor', 'Producto/Servicio', 
                'Método Pago', 'Frecuencia', 'Valor', 'Fecha Vencimiento',
                'Usuario Responsable', 'Cargo Responsable', 'Observaciones', 'Estado', 'Baja'
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Datos
            hoy = date.today()
            for row_num, lic in enumerate(licencias, 2):
                # Convertir fecha de string ISO a date si es necesario
                fecha_venc = None
                if lic.get('fechaVencimiento'):
                    if isinstance(lic['fechaVencimiento'], str):
                        fecha_venc = datetime.fromisoformat(lic['fechaVencimiento']).date()
                    else:
                        fecha_venc = lic['fechaVencimiento']
                
                # Calcular estado
                estado_lic = "Vigente"
                if fecha_venc:
                    dias_restantes = (fecha_venc - hoy).days
                    if dias_restantes < 0:
                        estado_lic = "Vencida"
                    elif dias_restantes <= 8:
                        estado_lic = "Crítica"
                    elif dias_restantes <= 30:
                        estado_lic = "Próxima a vencer"
                
                # Escribir datos
                data_row = [
                    lic.get('id', ''),
                    lic.get('tipoServicio', ''),
                    lic.get('proveedor', ''),
                    lic.get('producto', ''),
                    lic.get('metodoPago', ''),
                    lic.get('frecuencia', ''),
                    lic.get('valor', ''),
                    fecha_venc.strftime('%Y-%m-%d') if fecha_venc else '',
                    lic.get('responsable', {}).get('nombre', '') if isinstance(lic.get('responsable'), dict) else '',
                    lic.get('responsable', {}).get('cargo', '') if isinstance(lic.get('responsable'), dict) else '',
                    lic.get('observaciones', ''),
                    estado_lic,
                    'Sí' if lic.get('baja') else 'No'
                ]
                
                for col_num, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
            
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Guardar en BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
            
        except Exception as e:
            print(f"Error generando Excel: {e}")
            print(f"Traceback: {traceback.format_exc()}")
            raise CustomException(f"Error generando Excel: {str(e)}")
    
    # ===================================================
    # MÉTODOS PARA REVISIONES GENERALES
    # ===================================================

    def obtener_tipos_revision(self):
        """Obtiene todos los tipos de revisión disponibles"""
        try:
            tipos = self.querys.obtener_tipos_revision()
            return self.tools.output(200, "Tipos de revisión obtenidos exitosamente.", tipos)
        except CustomException as ce:
            return self.tools.output(500, str(ce), [])
        except Exception as e:
            print(f"Error obteniendo tipos de revisión: {e}")
            return self.tools.output(500, "Error obteniendo tipos de revisión.", [])

    def crear_revision(self, data):
        """Crea una nueva revisión general"""
        try:
            # Validar datos requeridos
            if not data.get('fecha'):
                return self.tools.output(400, "La fecha es requerida.", {})
            if not data.get('tipo_revision_id'):
                return self.tools.output(400, "El tipo de revisión es requerido.", {})
            if not data.get('usuario'):
                return self.tools.output(400, "El usuario es requerido.", {})
            
            revision = self.querys.crear_revision(data)
            return self.tools.output(201, "Revisión creada exitosamente.", revision)
        except CustomException as ce:
            return self.tools.output(400, str(ce), {})
        except Exception as e:
            print(f"Error creando revisión: {e}")
            return self.tools.output(500, "Error creando revisión.", {})

    def obtener_revisiones(self, page=1, per_page=5):
        """Obtiene todas las revisiones con paginación"""
        try:
            resultado = self.querys.obtener_revisiones(page, per_page)
            return self.tools.output(200, "Revisiones obtenidas exitosamente.", resultado)
        except CustomException as ce:
            return self.tools.output(500, str(ce), {})
        except Exception as e:
            print(f"Error obteniendo revisiones: {e}")
            return self.tools.output(500, "Error obteniendo revisiones.", {})

    # ===================================================
    # MÉTODOS PARA VERSIONES DEL CONTROL DE LICENCIAS
    # ===================================================

    def crear_version(self, data):
        """Crea una nueva versión del control de licencias"""
        try:
            # Validar datos requeridos
            if not data.get('fecha'):
                return self.tools.output(400, "La fecha es requerida.", {})
            if not data.get('version'):
                return self.tools.output(400, "La versión es requerida.", {})
            
            version = self.querys.crear_version(data)
            return self.tools.output(201, "Versión creada exitosamente.", version)
        except CustomException as ce:
            return self.tools.output(400, str(ce), {})
        except Exception as e:
            print(f"Error creando versión: {e}")
            return self.tools.output(500, "Error creando versión.", {})

    def obtener_versiones(self, page=1, per_page=5):
        """Obtiene todas las versiones con paginación"""
        try:
            resultado = self.querys.obtener_versiones(page, per_page)
            return self.tools.output(200, "Versiones obtenidas exitosamente.", resultado)
        except CustomException as ce:
            return self.tools.output(500, str(ce), {})
        except Exception as e:
            print(f"Error obteniendo versiones: {e}")
            return self.tools.output(500, "Error obteniendo versiones.", {})

    def eliminar_version(self, data: dict):
        """Elimina una versión"""
        try:
            version_id = data.get('version_id')
            
            if not version_id:
                raise CustomException("No se proporcionó version_id")
            
            resultado = self.querys.eliminar_version(version_id)
            return self.tools.output(200, resultado["message"], resultado)
        except CustomException as ce:
            return self.tools.output(404, str(ce), {})
        except Exception as e:
            print(f"Error eliminando versión: {e}")
            return self.tools.output(500, "Error eliminando versión.", {})
